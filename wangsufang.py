# -*- coding: utf-8 -*-
# @Time : 2022/3/2 18:47
# @Author : su.xinhai
# @Email : su.xinhai@mech-mind.net

import openpyxl
import xlrd
import os

dir = r"temp"
all_data = []
ec_file = r"C:\Users\mech-mind\PycharmProjects\StudyPy\1.xlsx"
sheet_name = "sheet1"
# 需要取数据的行
row_nums = [30, 31, 35, 9, 23, 16, 17, 20]
# 需要取数据的列
cell_num = 3


def create_ex():
    if os.path.exists(ec_file):
        return
    # 创建一个工作簿
    wb = openpyxl.Workbook()
    # 创建一个test_case的sheet表单
    wb.create_sheet(sheet_name)
    # 保存为一个xlsx格式的文件
    wb.save(ec_file)
    wb.close()


def insert_data(data):
    wb = openpyxl.load_workbook(ec_file)
    sh = wb[sheet_name]
    row = sh.max_row + 1
    for i in data:
        col = 1
        for j in i:
            sh.cell(row=row, column=col, value=j)
            col += 1
        row += 1

    wb.save(ec_file)
    wb.close()


def get_all_data(cur_file):
    global all_data
    all_file = os.listdir(cur_file)
    for file in all_file:
        file_path = os.path.join(cur_file, file)
        if not os.path.isdir(file_path):
            all_data.append(get_one_table(file_path))
        else:
            get_all_data(file_path)
    return all_data


def get_one_table(ex_file):
    data = []
    xls_file = xlrd.open_workbook(ex_file)
    table = xls_file.sheet_by_name('Sheet1')
    for i in row_nums:
        value1 = table.cell(i - 1, cell_num - 1).value
        data.append(value1)
    return data


if __name__ == '__main__':
    create_ex()
    insert_data(get_all_data(dir))
