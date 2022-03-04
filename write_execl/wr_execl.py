# -*- coding: utf-8 -*-
# @Time : 2022/2/28 14:18
# @Author : su.xinhai
# @Email : su.xinhai@mech-mind.net
import openpyxl
import os
import time

# xlsx文件路径
ex_file = r"C:\Users\mech-mind\PycharmProjects\StudyPy\write_execl\pose_accuracy.xlsx"
# vin码文件夹路径
vin_dir = r"C:\Users\mech-mind\PycharmProjects\StudyPy\write_execl"
# sheet名称
sheet_name = "sheet1"
# 车门名称
STATION = 1


def create_ex():
    if os.path.exists(ex_file):
        return
    # 创建一个工作簿
    wb = openpyxl.Workbook()
    # 创建一个test_case的sheet表单
    wb.create_sheet(sheet_name)
    # 保存为一个xlsx格式的文件
    wb.save(ex_file)
    wb.close()
    create_title()


def create_title():
    wb = openpyxl.load_workbook(ex_file)
    # 第二步：选取表单
    sh = wb[sheet_name]
    # 第三步：读取数据
    # 参数 row:行  column：列
    sh.cell(row=1, column=1, value='Time')
    sh.cell(row=1, column=2, value='vin码')
    sh.cell(row=1, column=3, value='工作站')
    j = 3
    for i in range(15):
        sh.cell(row=1, column=j + 1, value='X')
        sh.cell(row=1, column=j + 2, value='Y')
        sh.cell(row=1, column=j + 3, value='Z')
        j += 3
    wb.save(ex_file)
    wb.close()


def insert_data(pose):
    wb = openpyxl.load_workbook(ex_file)
    # 第二步：选取表单
    sh = wb[sheet_name]
    # 第三步：读取数据
    # 参数 row:行  column：列
    vin_code = get_vin()
    row = sh.max_row + 1
    sh.cell(row=row, column=1, value=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    sh.cell(row=row, column=2, value=vin_code)
    sh.cell(row=row, column=3, value=STATION)
    j = 3
    for i in range(len(pose)):
        sh.cell(row=row, column=j + 1, value=pose[i][0])
        sh.cell(row=row, column=j + 2, value=pose[i][1])
        sh.cell(row=row, column=j + 3, value=pose[i][2])
        j += 3
    for i in range(15 - len(pose)):
        sh.cell(row=row, column=j + 1, value=0)
        sh.cell(row=row, column=j + 2, value=0)
        sh.cell(row=row, column=j + 3, value=0)
        j += 3

    wb.save(ex_file)
    wb.close()


def get_vin():
    if not os.listdir(vin_dir):
        return ""
    for file in os.listdir(vin_dir):
        return file.split(".")[0]


if __name__ == '__main__':
    create_ex()
    pose = [[2.03697390471856, 0.9161494004171391, 1.096746932996529, 0.7095865578101006, 2.6020852139652106e-18,
             5.204170427930421e-18, -0.7046182774915888],
            [3.03697390471856, 0.9161494004171391, 1.096746932996529, 0.7095865578101006, 2.6020852139652106e-18,
             5.204170427930421e-18, -0.7046182774915888],
            [4.03697390471856, 0.9161494004171391, 1.096746932996529, 0.7095865578101006, 2.6020852139652106e-18,
             5.204170427930421e-18, -0.7046182774915888],
            [5.03697390471856, 0.9161494004171391, 1.096746932996529, 0.7095865578101006, 2.6020852139652106e-18,
             5.204170427930421e-18, -0.7046182774915888],
            [6.03697390471856, 0.9161494004171391, 1.096746932996529, 0.7095865578101006, 2.6020852139652106e-18,
             5.204170427930421e-18, -0.7046182774915888]]
    insert_data(pose)
import threading
threading.Thread(target=get_vin,args=())